"""
Management command: cargar_catalogo
Carga el catalogo completo desde catalogo_productos.xlsx.
Copia fotos disponibles desde static/img/Fotos Productos/ a media/products/.

Uso:
    python manage.py cargar_catalogo
    python manage.py cargar_catalogo --limpiar      # borra productos existentes primero
    python manage.py cargar_catalogo --dry-run      # no guarda nada, solo muestra resumen
"""

import shutil
from pathlib import Path

from django.conf import settings
from django.core.management.base import BaseCommand

from apps.productos.models import Categoria, Producto


FOTO_DIRS = ['Tabacos', 'Accesorios']
WEBP_CALIDAD = 82
WEBP_MAX_DIM = 800


def buscar_foto(nombre_archivo, base_fotos):
    """Busca la foto en la raíz y subcarpetas de Fotos Productos/.
    Acepta cualquier extensión de imagen equivalente."""
    stem = Path(nombre_archivo).stem
    extensiones = ['.webp', '.jpg', '.jpeg', '.png', '.JPG', '.JPEG', '.PNG']

    # Directorios donde buscar: raíz + subcarpetas
    dirs_busqueda = [base_fotos] + [base_fotos / sub for sub in FOTO_DIRS]

    for directorio in dirs_busqueda:
        for ext in extensiones:
            ruta = directorio / (stem + ext)
            if ruta.exists():
                return ruta
    return None


def copiar_como_webp(origen: Path, destino_dir: Path) -> str:
    """Copia la imagen convirtiéndola a WebP. Retorna el nombre final del archivo."""
    from PIL import Image

    nombre_webp = origen.stem + '.webp'
    destino = destino_dir / nombre_webp

    if destino.exists():
        return nombre_webp

    with Image.open(origen) as img:
        if img.mode in ('RGBA', 'LA', 'P'):
            img = img.convert('RGBA')
            fondo = Image.new('RGBA', img.size, (255, 255, 255, 255))
            fondo.paste(img, mask=img.split()[3])
            img = fondo.convert('RGB')
        elif img.mode != 'RGB':
            img = img.convert('RGB')
        img.thumbnail((WEBP_MAX_DIM, WEBP_MAX_DIM), Image.LANCZOS)
        img.save(destino, 'WEBP', quality=WEBP_CALIDAD, method=6)

    return nombre_webp


class Command(BaseCommand):
    help = "Carga el catalogo desde catalogo_productos.xlsx"

    def add_arguments(self, parser):
        parser.add_argument(
            "--limpiar",
            action="store_true",
            help="Elimina todos los productos existentes antes de cargar",
        )
        parser.add_argument(
            "--dry-run",
            action="store_true",
            dest="dry_run",
            help="Muestra lo que se haria sin guardar nada",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
        except ImportError:
            self.stderr.write(self.style.ERROR(
                "openpyxl no instalado. Ejecuta: pip install openpyxl"
            ))
            return

        dry_run = options["dry_run"]
        limpiar = options["limpiar"]

        base_dir = Path(settings.BASE_DIR)
        excel_path = base_dir / "catalogo_productos.xlsx"
        base_fotos = base_dir / "static" / "img" / "Fotos Productos"
        media_products = Path(settings.MEDIA_ROOT) / "products"

        if not excel_path.exists():
            self.stderr.write(self.style.ERROR(f"No se encontro: {excel_path}"))
            return

        if not dry_run:
            media_products.mkdir(parents=True, exist_ok=True)

        self.stdout.write("Leyendo catalogo_productos.xlsx ...")
        wb = openpyxl.load_workbook(str(excel_path))
        ws = wb.active

        headers = []
        for c in range(1, ws.max_column + 1):
            val = ws.cell(1, c).value
            if val:
                headers.append(val)
            else:
                break

        def col(row, field):
            try:
                idx = headers.index(field) + 1
                return ws.cell(row, idx).value
            except ValueError:
                return None

        if limpiar and not dry_run:
            n = Producto.objects.count()
            Producto.objects.all().delete()
            Categoria.objects.all().delete()
            self.stdout.write(self.style.WARNING(
                f"Eliminados {n} productos y todas las categorias."
            ))

        prods_creados = 0
        prods_actualizados = 0
        fotos_copiadas = 0
        fotos_ya_existian = 0
        fotos_faltantes = 0
        fotos_sin_campo = 0

        for row in range(2, ws.max_row + 1):
            nombre = col(row, "nombre")
            if not nombre:
                continue

            seccion = col(row, "seccion") or ""
            categoria_nombre = col(row, "categoria") or "Sin categoria"
            marca = col(row, "marca") or ""
            codigo = col(row, "codigo") or None
            try:
                peso_gramos = int(col(row, "peso_gramos") or 0) or None
            except (ValueError, TypeError):
                peso_gramos = None
            dimensiones = col(row, "dimensiones") or ""
            procedencia = col(row, "procedencia") or ""
            intensidad = col(row, "intensidad") or ""
            descripcion = col(row, "descripcion") or ""
            precio_unidad = col(row, "precio_unidad") or 0
            stock = col(row, "stock") or 0
            activo_excel = str(col(row, "activo") or "SI").upper() == "SI"
            activo = False  # se actualiza abajo según foto + columna activo
            imagen_nombre  = col(row, "imagen")   or ""
            imagen_nombre_2 = col(row, "imagen_2") or ""
            imagen_nombre_3 = col(row, "imagen_3") or ""
            foto_disponible = str(col(row, "foto_disponible") or "NO").upper() == "SI"

            if intensidad not in ("Suave", "Medio", "Intenso"):
                intensidad = ""

            def procesar_imagen_extra(nombre_img):
                """Copia/convierte una imagen adicional. Retorna el campo o ''."""
                if not nombre_img:
                    return ""
                origen_extra = buscar_foto(nombre_img, base_fotos)
                if origen_extra:
                    if not dry_run:
                        nf = copiar_como_webp(origen_extra, media_products)
                        dest = media_products / nf
                        if dest.exists():
                            from apps.productos.watermark import aplicar_sello
                            aplicar_sello(str(dest))
                        return f"products/{nf}"
                    else:
                        return f"products/{Path(nombre_img).stem}.webp"
                # Si ya está en media directamente
                webp_nombre = Path(nombre_img).stem + '.webp'
                if (media_products / webp_nombre).exists():
                    return f"products/{webp_nombre}"
                if (media_products / nombre_img).exists():
                    return f"products/{nombre_img}"
                return ""

            imagen_field = ""
            if foto_disponible and imagen_nombre:
                # 1. Buscar en static (originales)
                origen = buscar_foto(imagen_nombre, base_fotos)
                if origen:
                    if not dry_run:
                        nombre_final = copiar_como_webp(origen, media_products)
                        destino = media_products / nombre_final
                        es_nueva = destino.stat().st_size > 0
                        if es_nueva:
                            from apps.productos.watermark import aplicar_sello
                            aplicar_sello(str(destino))
                            fotos_copiadas += 1
                        else:
                            fotos_ya_existian += 1
                    else:
                        nombre_final = Path(imagen_nombre).stem + '.webp'
                    imagen_field = f"products/{nombre_final}"
                    activo = activo_excel
                else:
                    # 2. Buscar directamente en media/products/ (ya convertidas)
                    webp_nombre = Path(imagen_nombre).stem + '.webp'
                    if (media_products / webp_nombre).exists():
                        imagen_field = f"products/{webp_nombre}"
                        activo = activo_excel
                        fotos_ya_existian += 1
                    elif (media_products / imagen_nombre).exists():
                        imagen_field = f"products/{imagen_nombre}"
                        activo = activo_excel
                        fotos_ya_existian += 1
                    else:
                        fotos_faltantes += 1
                        self.stdout.write(
                            self.style.WARNING(f"  Foto no encontrada: {imagen_nombre} ({nombre})")
                        )
            else:
                fotos_sin_campo += 1

            if dry_run:
                if imagen_field:
                    prods_creados += 1
                else:
                    prods_creados += 1
                continue

            categoria, _ = Categoria.objects.get_or_create(
                nombre=categoria_nombre,
                defaults={"descripcion": ""},
            )

            imagen_field_2 = procesar_imagen_extra(imagen_nombre_2)
            imagen_field_3 = procesar_imagen_extra(imagen_nombre_3)

            defaults = {
                "nombre": nombre,
                "seccion": seccion,
                "marca": marca,
                "categoria": categoria,
                "peso_gramos": peso_gramos,
                "dimensiones": dimensiones,
                "procedencia": procedencia,
                "intensidad": intensidad,
                "descripcion": descripcion,
                "precio_unidad": precio_unidad,
                "stock": stock,
                "activo": activo,
                "destacado": False,
            }
            if imagen_field:
                defaults["imagen"] = imagen_field
            if imagen_field_2:
                defaults["imagen_2"] = imagen_field_2
            if imagen_field_3:
                defaults["imagen_3"] = imagen_field_3

            if codigo:
                prod, created = Producto.objects.update_or_create(
                    codigo=codigo,
                    defaults=defaults,
                )
            else:
                prod, created = Producto.objects.update_or_create(
                    nombre=nombre,
                    marca=marca,
                    defaults=defaults,
                )

            if created:
                prods_creados += 1
            else:
                prods_actualizados += 1

        self.stdout.write("")
        if dry_run:
            self.stdout.write(self.style.WARNING("--- DRY RUN (nada guardado) ---"))
            self.stdout.write(f"  Productos a cargar     : {prods_creados}")
            self.stdout.write(f"  Con foto en disco      : {prods_creados - fotos_faltantes - fotos_sin_campo}")
            self.stdout.write(f"  Sin foto (foto_disp=NO): {fotos_sin_campo}")
            self.stdout.write(f"  Foto marcada pero falta: {fotos_faltantes}")
        else:
            self.stdout.write(self.style.SUCCESS("--- Catalogo cargado ---"))
            self.stdout.write(f"  Creados              : {prods_creados}")
            self.stdout.write(f"  Actualizados         : {prods_actualizados}")
            self.stdout.write(f"  Fotos copiadas       : {fotos_copiadas}")
            self.stdout.write(f"  Fotos ya existian    : {fotos_ya_existian}")
            self.stdout.write(f"  Fotos no encontradas : {fotos_faltantes}")
            self.stdout.write(f"  Productos sin foto   : {fotos_sin_campo}")
            self.stdout.write("")
            self.stdout.write(
                "Ahora ve al Admin de Django y marca algunos productos como 'destacado'"
            )
            self.stdout.write("para que aparezcan en la seccion de inicio.")
