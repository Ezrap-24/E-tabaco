"""
Comando: optimizar_imagenes
Convierte todas las imágenes de productos de JPG/PNG → WebP (800×800px, calidad 82).
Actualiza los registros en la DB y el Excel catalogo_productos.xlsx.

Uso:
    python manage.py optimizar_imagenes
    python manage.py optimizar_imagenes --dry-run       # solo muestra lo que haría
    python manage.py optimizar_imagenes --calidad 85    # ajustar calidad WebP (defecto: 82)
    python manage.py optimizar_imagenes --sin-excel     # no toca el Excel
"""
from pathlib import Path
from django.conf import settings
from django.core.management.base import BaseCommand
from apps.productos.models import Producto


MAX_DIM = 800
FORMATOS_ORIGEN = {'.jpg', '.jpeg', '.png', '.JPG', '.JPEG', '.PNG'}


def convertir_a_webp(ruta_origen: Path, calidad: int) -> Path:
    """Convierte una imagen a WebP redimensionada. Retorna la ruta del nuevo archivo."""
    from PIL import Image

    ruta_destino = ruta_origen.with_suffix('.webp')

    with Image.open(ruta_origen) as img:
        # Convertir a RGB (por si tiene canal alfa o es CMYK)
        if img.mode in ('RGBA', 'LA', 'P'):
            img = img.convert('RGBA')
            fondo = Image.new('RGBA', img.size, (255, 255, 255, 255))
            fondo.paste(img, mask=img.split()[3])
            img = fondo.convert('RGB')
        elif img.mode != 'RGB':
            img = img.convert('RGB')

        # Redimensionar manteniendo proporción
        img.thumbnail((MAX_DIM, MAX_DIM), Image.LANCZOS)
        img.save(ruta_destino, 'WEBP', quality=calidad, method=6)

    return ruta_destino


class Command(BaseCommand):
    help = 'Convierte imágenes de productos a WebP y actualiza DB y Excel.'

    def add_arguments(self, parser):
        parser.add_argument('--dry-run', action='store_true', dest='dry_run',
                            help='Solo muestra lo que haría, sin guardar.')
        parser.add_argument('--calidad', type=int, default=82, dest='calidad',
                            help='Calidad WebP (1-100, defecto 82).')
        parser.add_argument('--sin-excel', action='store_true', dest='sin_excel',
                            help='No actualizar el archivo Excel.')

    def handle(self, *args, **options):
        dry_run = options['dry_run']
        calidad = options['calidad']
        sin_excel = options['sin_excel']

        media_products = Path(settings.MEDIA_ROOT) / 'products'

        if not media_products.exists():
            self.stderr.write(self.style.ERROR(f'No existe: {media_products}'))
            return

        # ── 1. Convertir imágenes en disco ────────────────────────
        self.stdout.write(self.style.HTTP_INFO('\n── Convirtiendo imágenes ──'))
        convertidos = []
        errores = []
        ya_webp = 0

        for archivo in sorted(media_products.iterdir()):
            if archivo.suffix.lower() == '.webp':
                ya_webp += 1
                continue
            if archivo.suffix not in FORMATOS_ORIGEN:
                continue

            try:
                tam_antes = archivo.stat().st_size
                if not dry_run:
                    ruta_webp = convertir_a_webp(archivo, calidad)
                    tam_despues = ruta_webp.stat().st_size
                    ahorro = (1 - tam_despues / tam_antes) * 100
                    convertidos.append((archivo.name, ruta_webp.name, tam_antes, tam_despues))
                    self.stdout.write(
                        f'  ✓ {archivo.name} → {ruta_webp.name} '
                        f'({tam_antes//1024}KB → {tam_despues//1024}KB, -{ahorro:.0f}%)'
                    )
                else:
                    convertidos.append((archivo.name, archivo.stem + '.webp', tam_antes, 0))
                    self.stdout.write(f'  [dry] {archivo.name} → {archivo.stem}.webp')
            except Exception as e:
                errores.append((archivo.name, str(e)))
                self.stdout.write(self.style.ERROR(f'  ✗ {archivo.name}: {e}'))

        # ── 2. Actualizar DB ──────────────────────────────────────
        self.stdout.write(self.style.HTTP_INFO('\n── Actualizando base de datos ──'))
        actualizados_db = 0

        for nombre_orig, nombre_webp, _, _ in convertidos:
            campo_orig = f'products/{nombre_orig}'
        for nombre_orig, nombre_webp, _, _ in convertidos:
            campo_orig = f'products/{nombre_orig}'
            campo_webp = f'products/{nombre_webp}'
            qs = Producto.objects.filter(imagen=campo_orig)
            count = qs.count()
            if count:
                if not dry_run:
                    qs.update(imagen=campo_webp)
                self.stdout.write(f'  ✓ DB: {campo_orig} → {campo_webp} ({count} producto/s)')
                actualizados_db += count

        # ── 3. Actualizar Excel ───────────────────────────────────
        actualizados_excel = 0
        if not sin_excel:
            base_dir = Path(settings.BASE_DIR)
            excel_path = base_dir / 'catalogo_productos.xlsx'

            if not excel_path.exists():
                self.stdout.write(self.style.WARNING(
                    f'\n── Excel no encontrado en {excel_path} — omitiendo.'
                ))
            else:
                self.stdout.write(self.style.HTTP_INFO('\n── Actualizando Excel ──'))
                try:
                    import openpyxl
                    wb = openpyxl.load_workbook(str(excel_path))
                    ws = wb.active

                    # Encontrar columna "imagen"
                    col_imagen = None
                    for c in range(1, ws.max_column + 1):
                        if ws.cell(1, c).value == 'imagen':
                            col_imagen = c
                            break

                    if col_imagen is None:
                        self.stdout.write(self.style.WARNING('  Columna "imagen" no encontrada en el Excel.'))
                    else:
                        nombres_convertidos = {n_orig: n_webp for n_orig, n_webp, _, _ in convertidos}
                        for row in range(2, ws.max_row + 1):
                            val = ws.cell(row, col_imagen).value
                            if val and val in nombres_convertidos:
                                nuevo = nombres_convertidos[val]
                                if not dry_run:
                                    ws.cell(row, col_imagen).value = nuevo
                                self.stdout.write(f'  ✓ Excel fila {row}: {val} → {nuevo}')
                                actualizados_excel += 1

                        if not dry_run:
                            wb.save(str(excel_path))
                            self.stdout.write(self.style.SUCCESS(f'  Excel guardado: {excel_path}'))
                except Exception as e:
                    self.stdout.write(self.style.ERROR(f'  Error actualizando Excel: {e}'))

        # ── 4. Eliminar originales ────────────────────────────────
        if not dry_run and convertidos:
            self.stdout.write(self.style.HTTP_INFO('\n── Eliminando originales ──'))
            for nombre_orig, _, _, _ in convertidos:
                ruta_orig = media_products / nombre_orig
                if ruta_orig.exists():
                    ruta_orig.unlink()
                    self.stdout.write(f'  🗑  {nombre_orig}')

        # ── Resumen ───────────────────────────────────────────────
        if convertidos and not dry_run:
            total_antes = sum(t for _, _, t, _ in convertidos)
            total_despues = sum(t for _, _, _, t in convertidos)
            ahorro_total = (1 - total_despues / total_antes) * 100 if total_antes else 0
            ahorro_mb = (total_antes - total_despues) / (1024 * 1024)
        else:
            total_antes = sum(t for _, _, t, _ in convertidos)
            ahorro_total = ahorro_mb = 0

        self.stdout.write('')
        prefijo = '[DRY RUN] ' if dry_run else ''
        self.stdout.write(self.style.SUCCESS(f'{prefijo}── Resumen ──'))
        self.stdout.write(f'  Imágenes convertidas : {len(convertidos)}')
        self.stdout.write(f'  Ya eran WebP         : {ya_webp}')
        self.stdout.write(f'  Errores              : {len(errores)}')
        self.stdout.write(f'  Productos actualizados DB   : {actualizados_db}')
        self.stdout.write(f'  Filas actualizadas Excel    : {actualizados_excel}')
        if not dry_run and convertidos:
            self.stdout.write(
                self.style.SUCCESS(
                    f'  Espacio liberado     : ~{ahorro_mb:.0f} MB ({ahorro_total:.0f}% de reducción)'
                )
            )
