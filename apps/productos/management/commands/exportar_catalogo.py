"""
Management command: exportar_catalogo
Exporta todos los productos a un archivo Excel preservando todos los datos.

Uso:
    python manage.py exportar_catalogo                    # exporta a catalogo-export.xlsx
    python manage.py exportar_catalogo --output archivo.xlsx
"""

import os
from pathlib import Path
from django.conf import settings
from django.core.management.base import BaseCommand
from apps.productos.models import Producto


class Command(BaseCommand):
    help = "Exporta todos los productos a Excel con todos sus datos"

    def add_arguments(self, parser):
        parser.add_argument(
            "--output",
            type=str,
            default="catalogo-export.xlsx",
            help="Nombre del archivo de salida (default: catalogo-export.xlsx)",
        )

    def handle(self, *args, **options):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            self.stderr.write(self.style.ERROR(
                "openpyxl no instalado. Ejecuta: pip install openpyxl"
            ))
            return

        output_filename = options["output"]

        # Crear el archivo Excel
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Catalogo"

        # Definir headers exactamente como en cargar_catalogo.py
        headers = [
            "nombre",
            "marca",
            "codigo",
            "categoria",
            "peso_gramos",
            "procedencia",
            "intensidad",
            "descripcion",
            "precio_unidad",
            "precio_mayor",
            "cantidad_minima_mayor",
            "stock",
            "activo",
            "imagen",
            "foto_disponible",
        ]

        # Escribir headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

        # Información media
        base_dir = Path(settings.BASE_DIR)
        media_products = Path(settings.MEDIA_ROOT) / "products"

        # Obtener todos los productos
        productos = Producto.objects.all().order_by('marca', 'nombre')

        self.stdout.write(f"Exportando {productos.count()} productos...")

        row = 3
        for prod in productos:
            # Determinar si la foto está disponible
            foto_disponible = "NO"
            if prod.imagen:
                # Obtener nombre de archivo de la imagen
                imagen_nombre = str(prod.imagen).replace('products/', '')
                imagen_path = media_products / imagen_nombre
                if imagen_path.exists():
                    foto_disponible = "SI"

            # Escribir datos del producto
            ws.cell(row=row, column=1, value=prod.nombre)
            ws.cell(row=row, column=2, value=prod.marca)
            ws.cell(row=row, column=3, value=prod.codigo or "")
            ws.cell(row=row, column=4, value=prod.categoria.nombre if prod.categoria else "")
            ws.cell(row=row, column=5, value=prod.peso_gramos or "")
            ws.cell(row=row, column=6, value=prod.procedencia or "")
            ws.cell(row=row, column=7, value=prod.intensidad or "")
            ws.cell(row=row, column=8, value=prod.descripcion or "")
            ws.cell(row=row, column=9, value=float(prod.precio_unidad) if prod.precio_unidad else 0)
            ws.cell(row=row, column=10, value=float(prod.precio_mayor) if prod.precio_mayor else "")
            ws.cell(row=row, column=11, value=prod.cantidad_minima_mayor or 3)
            ws.cell(row=row, column=12, value=prod.stock or 0)
            ws.cell(row=row, column=13, value="SI" if prod.activo else "NO")
            ws.cell(row=row, column=14, value=str(prod.imagen) if prod.imagen else "")
            ws.cell(row=row, column=15, value=foto_disponible)

            row += 1

        # Ajustar ancho de columnas
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['H'].width = 40

        # Guardar
        output_path = base_dir / output_filename
        wb.save(str(output_path))

        self.stdout.write(self.style.SUCCESS(f"\n✓ Exportado: {output_path}"))
        self.stdout.write(f"✓ Total productos: {productos.count()}")
        self.stdout.write(f"✓ Columnas: {', '.join(headers)}")
