"""
Convierte productos_export.csv a productos_export.xlsx
Uso: python crear_excel_desde_csv.py
"""
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

INPUT = 'productos_export.csv'
OUTPUT = 'productos_export.xlsx'

wb = Workbook()
ws = wb.active
ws.title = 'Productos'

header_font = Font(bold=True, color='FFFFFF')
header_fill = PatternFill(fill_type='solid', fgColor='2D6A4F')

with open(INPUT, newline='', encoding='utf-8') as f:
    reader = csv.reader(f)
    for i, row in enumerate(reader):
        ws.append(row)
        if i == 0:
            for cell in ws[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')

for col in ws.columns:
    max_len = max((len(str(cell.value or '')) for cell in col), default=10)
    ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 50)

wb.save(OUTPUT)
print(f'Excel guardado -> {OUTPUT}')
